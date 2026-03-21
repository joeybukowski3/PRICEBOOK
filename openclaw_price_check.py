"""
OPENCLAW Price Check - complete rewrite
Two email tables: vs listed price + vs previous run
SKU column with BB links, download link in footer
"""
import requests, re, time, os, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

BESTBUY_API_KEY = 'twmOxg13zZdgLikzAfGATIXl'
EXCEL_FILE      = 'GITHUBOPENCLAWPRICECHECK.xlsx'
TODAY_LABEL     = f"{date.today().month}/{date.today().day}"
REPO_URL        = "https://github.com/joeybukowski3/PRICEBOOK"
EXCEL_URL       = "https://github.com/joeybukowski3/PRICEBOOK/raw/main/GITHUBOPENCLAWPRICECHECK.xlsx"

thin    = Side(style='thin', color='D1D5DB')
T       = Border(left=thin, right=thin, top=thin, bottom=thin)
TEAL_L  = "E0FAF5"; GREEN_BG = "DCFCE7"; GREEN_FG = "15803D"
AMBER_L = "FFFBEB"; AMBER    = "F59E0B"
LGRAY   = "F3F4F6"; DGRAY    = "374151"
NAVY    = "0D1B2A"; WHITE    = "FFFFFF"
RED_BG  = "FEE2E2"; RED_FG   = "DC2626"

def check_sku(sku):
    url    = f'https://api.bestbuy.com/v1/products/{sku}.json'
    params = {'apiKey': BESTBUY_API_KEY,
              'show':   'sku,name,salePrice,regularPrice,onlineAvailability'}
    try:
        r = requests.get(url, params=params, timeout=10)
        if r.status_code == 200:
            d = r.json()
            return (d.get('salePrice') or d.get('regularPrice'),
                    d.get('onlineAvailability', False), d.get('name', ''))
        elif r.status_code == 404: return None, False, 'NOT FOUND'
        elif r.status_code == 403: return None, False, 'RATE LIMITED'
        else: return None, False, f'API ERROR {r.status_code}'
    except Exception as e:
        return None, False, str(e)[:60]

def clean_sku(raw):
    if raw is None: return None
    s = str(raw).replace('.0', '').strip()
    if s.isdigit(): return s
    m = re.search(r'skuId=(\d+)|/(\d{7,8})\.p', s)
    return (m.group(1) or m.group(2)) if m else None

def send_email(checked, unchanged, n_listed, oos, not_found,
               errors, vs_listed, vs_prev, today):
    resend_key   = os.environ.get('RESEND_API_KEY', '')
    notify_email = os.environ.get('NOTIFY_EMAIL', '')
    if not resend_key or not notify_email:
        print("No RESEND_API_KEY or NOTIFY_EMAIL — skipping email")
        return

    def table_rows(data):
        # data: list of (itype, name, sku, price_a, price_b, pct)
        if not data:
            return '<tr><td colspan="6" style="padding:14px;text-align:center;color:#6B7280;font-size:12px">&#x2705; No changes</td></tr>'
        html = ""
        for itype, name, sku, pa, pb, pct in sorted(data, key=lambda x: abs(x[5]), reverse=True):
            color  = "#DC2626" if pb > pa else "#15803D"
            bg     = "#FFF7ED" if pb > pa else "#DCFCE7"
            arrow  = "&#8599;" if pb > pa else "&#8600;"
            sku_s  = str(sku) if sku else "—"
            link   = f"https://www.bestbuy.com/site/-.p?skuId={sku_s}" if sku_s.isdigit() else "#"
            html += f"""<tr style="background:{bg}">
<td style="padding:6px 10px;font-size:12px;color:#374151">{itype}</td>
<td style="padding:6px 10px;font-size:12px;color:#374151">{(name or '')[:45]}</td>
<td style="padding:6px 10px;font-size:12px;text-align:center"><a href="{link}" style="color:#2563EB;font-size:11px">{sku_s}</a></td>
<td style="padding:6px 10px;font-size:12px;text-align:center">${pa:.2f}</td>
<td style="padding:6px 10px;font-size:12px;text-align:center;font-weight:bold;color:{color}">${pb:.2f}</td>
<td style="padding:6px 10px;font-size:12px;text-align:center;font-weight:bold;color:{color}">{arrow} {pct:+.1f}%</td>
</tr>"""
        return html

    def make_table(title, h1, h2, rows):
        return f"""<h3 style="margin:20px 0 6px;font-size:12px;font-weight:bold;color:#1F2937">{title}</h3>
<table style="width:100%;border-collapse:collapse;background:white;border:1px solid #E5E7EB;border-radius:6px;overflow:hidden">
<thead><tr style="background:#1F2937">
<th style="padding:8px 10px;text-align:left;color:white;font-size:11px">Type</th>
<th style="padding:8px 10px;text-align:left;color:white;font-size:11px">Item</th>
<th style="padding:8px 10px;text-align:center;color:white;font-size:11px">SKU</th>
<th style="padding:8px 10px;text-align:center;color:white;font-size:11px">{h1}</th>
<th style="padding:8px 10px;text-align:center;color:white;font-size:11px">{h2}</th>
<th style="padding:8px 10px;text-align:center;color:white;font-size:11px">Change</th>
</tr></thead>
<tbody>{rows}</tbody></table>"""

    t1 = make_table("CHANGES VS LISTED PRICE", "Listed", "Today", table_rows(vs_listed))
    t2 = make_table("CHANGES SINCE LAST RUN", "Prev Run", "Today", table_rows(vs_prev))
    n_prev = len(vs_prev)

    html = f"""<div style="font-family:Arial,sans-serif;max-width:720px;margin:0 auto">
<div style="background:#0D1B2A;padding:20px;border-radius:8px 8px 0 0">
<h2 style="color:white;margin:0;font-size:18px">&#x1F4CA; OPENCLAW Daily Price Check</h2>
<p style="color:#9CA3AF;margin:4px 0 0;font-size:13px">{today}</p>
</div>
<div style="background:#F9FAFB;padding:20px;border:1px solid #E5E7EB;border-top:none;border-radius:0 0 8px 8px">
<div style="display:flex;gap:10px;margin-bottom:20px">
<div style="flex:1;background:white;border:1px solid #E5E7EB;border-radius:6px;padding:10px;text-align:center">
<div style="font-size:20px;font-weight:bold;color:#0D1B2A">{checked}</div>
<div style="font-size:11px;color:#6B7280">Checked</div></div>
<div style="flex:1;background:#DCFCE7;border:1px solid #BBF7D0;border-radius:6px;padding:10px;text-align:center">
<div style="font-size:20px;font-weight:bold;color:#15803D">{unchanged}</div>
<div style="font-size:11px;color:#15803D">Unchanged</div></div>
<div style="flex:1;background:{'#FFF7ED' if n_listed else '#F9FAFB'};border:1px solid {'#FED7AA' if n_listed else '#E5E7EB'};border-radius:6px;padding:10px;text-align:center">
<div style="font-size:20px;font-weight:bold;color:#{'EA580C' if n_listed else '6B7280'}">{n_listed}</div>
<div style="font-size:11px;color:#{'EA580C' if n_listed else '6B7280'}">vs Listed</div></div>
<div style="flex:1;background:{'#FFF7ED' if n_prev else '#F9FAFB'};border:1px solid {'#FED7AA' if n_prev else '#E5E7EB'};border-radius:6px;padding:10px;text-align:center">
<div style="font-size:20px;font-weight:bold;color:#{'EA580C' if n_prev else '6B7280'}">{n_prev}</div>
<div style="font-size:11px;color:#{'EA580C' if n_prev else '6B7280'}">New Changes</div></div>
<div style="flex:1;background:#FEE2E2;border:1px solid #FECACA;border-radius:6px;padding:10px;text-align:center">
<div style="font-size:20px;font-weight:bold;color:#DC2626">{not_found+oos}</div>
<div style="font-size:11px;color:#DC2626">OOS/404</div></div>
</div>
{t1}
{t2}
<p style="font-size:11px;color:#9CA3AF;margin-top:16px;text-align:center">
<a href="{EXCEL_URL}" style="color:#2563EB;font-weight:bold;text-decoration:none">&#x2B07; Download Latest Excel</a>
&nbsp;&nbsp;|&nbsp;&nbsp;
<a href="{REPO_URL}" style="color:#2563EB;text-decoration:none">View Repo</a>
</p></div></div>"""

    payload = {"from": "PriceCheck <onboarding@resend.dev>", "to": [notify_email],
               "subject": f"OPENCLAW {today} — {n_listed} vs listed | {n_prev} new changes",
               "html": html}
    try:
        resp = requests.post("https://api.resend.com/emails",
            headers={"Authorization": f"Bearer {resend_key}", "Content-Type": "application/json"},
            data=json.dumps(payload), timeout=15)
        print(f"Email {'sent' if resp.status_code in (200,201) else 'failed: '+str(resp.status_code)}")
    except Exception as e:
        print(f"Email error: {e}")

# ── LOAD WORKBOOK ─────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active

last_row = 1
for row in range(ws.max_row, 1, -1):
    if ws.cell(row=row, column=3).value:
        last_row = row
        break

# Find date columns, previous column, and write column
date_cols = []
next_col  = None
today_col = None
for col in range(1, ws.max_column + 2):
    val = ws.cell(row=1, column=col).value
    if val is None and next_col is None:
        next_col = col
    if val is not None:
        s = str(val)
        if ('/' in s and len(s) <= 5) or hasattr(val, 'month'):
            date_cols.append((col, s))
        if s == TODAY_LABEL:
            today_col = col

prev_col = None
for col, label in reversed(date_cols):
    if label != TODAY_LABEL:
        prev_col = col
        break

write_col = today_col if today_col else next_col
if not today_col:
    hdr = ws.cell(row=1, column=write_col, value=TODAY_LABEL)
    hdr.font = Font(color=WHITE, bold=True, size=9, name='Arial')
    hdr.fill = PatternFill('solid', start_color=NAVY)
    hdr.alignment = Alignment(horizontal='center', vertical='center')
    hdr.border = T
    ws.column_dimensions[hdr.column_letter].width = 12

print(f"\nOPENCLAW PRICE CHECK — {TODAY_LABEL}")
print(f"Write col: {write_col} | Prev col: {prev_col}")

checked = unchanged = changed = oos = not_found = errors = 0
vs_listed = []
vs_prev   = []

for row in range(2, last_row + 1):
    row_num   = ws.cell(row=row, column=3).value
    item_type = ws.cell(row=row, column=4).value
    name      = ws.cell(row=row, column=5).value
    sku_raw   = ws.cell(row=row, column=7).value
    listed    = ws.cell(row=row, column=9).value
    prev_val  = ws.cell(row=row, column=prev_col).value if prev_col else None

    if not row_num: continue

    sku = clean_sku(sku_raw)
    if not sku:
        cell = ws.cell(row=row, column=write_col, value='No SKU')
        cell.font = Font(color=DGRAY, size=9, name='Arial')
        cell.fill = PatternFill('solid', start_color=LGRAY)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = T
        errors += 1
        continue

    price, avail, bb_name = check_sku(sku)
    time.sleep(0.35)
    checked += 1

    if bb_name == 'RATE LIMITED':
        time.sleep(3)
        price, avail, bb_name = check_sku(sku)
        time.sleep(0.5)

    if price is None:
        cell = ws.cell(row=row, column=write_col, value='OOS' if not avail else bb_name[:12])
        cell.font = Font(color=RED_FG, bold=True, size=9, name='Arial')
        cell.fill = PatternFill('solid', start_color=RED_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = T
        not_found += 1
        print(f"  ❌ {sku} [{item_type}] — {bb_name}")
        continue

    if not avail:
        oos += 1

    # vs listed
    try:
        listed_f = float(str(listed).replace('$','').replace(',','')) if listed else None
    except: listed_f = None

    if listed_f and abs(price - listed_f) / listed_f > 0.01:
        pct = (price - listed_f) / listed_f * 100
        bg, fg = (AMBER_L, AMBER) if price > listed_f else (GREEN_BG, GREEN_FG)
        changed += 1
        vs_listed.append((item_type, name, sku, listed_f, price, pct))
        print(f"  {'📈' if price > listed_f else '📉'} {sku} listed:${listed_f:.2f} today:${price:.2f} ({pct:+.1f}%)")
    else:
        bg, fg = TEAL_L, "00897B"
        unchanged += 1

    # vs prev run
    try:
        prev_f = float(prev_val) if prev_val and isinstance(prev_val, (int, float)) else None
    except: prev_f = None
    if prev_f and abs(price - prev_f) / prev_f > 0.005:
        pct2 = (price - prev_f) / prev_f * 100
        vs_prev.append((item_type, name, sku, prev_f, price, pct2))
        print(f"  🔄 {sku} prev:${prev_f:.2f} today:${price:.2f} ({pct2:+.1f}%)")

    cell = ws.cell(row=row, column=write_col, value=price)
    cell.font = Font(color=fg, bold=(bg != TEAL_L), size=10, name='Arial')
    cell.fill = PatternFill('solid', start_color=bg)
    cell.number_format = '$#,##0.00'
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = T

wb.save(EXCEL_FILE)

print(f"\nDONE — checked:{checked} unchanged:{unchanged} vs_listed:{changed} new_changes:{len(vs_prev)} oos:{oos} not_found:{not_found}")
send_email(checked, unchanged, changed, oos, not_found, errors, vs_listed, vs_prev, TODAY_LABEL)
print(f"Saved: {EXCEL_FILE}")
