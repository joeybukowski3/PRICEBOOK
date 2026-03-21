"""
Daily BB Price Checker
Runs via GitHub Actions at 4 AM daily.
Updates PriceBook_Combined_Final.xlsx with a new date column.
Sends email summary via Resend.

Environment variables needed (set as GitHub Secrets):
  BESTBUY_API_KEY
  RESEND_API_KEY
  NOTIFY_EMAIL
"""
import os, re, time, pickle, requests, json
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BESTBUY_API_KEY = os.environ['BESTBUY_API_KEY']
RESEND_API_KEY  = os.environ['RESEND_API_KEY']
NOTIFY_EMAIL    = os.environ['NOTIFY_EMAIL']
EXCEL_FILE      = 'PriceBook_Combined_Final.xlsx'
ENTRIES_FILE    = 'combined_bb_final.pkl'
TODAY_LABEL     = f"{date.today().month}/{date.today().day}"

EXCLUDE = [
    "applecare","care+","geek squad","refurbished","open-box","renewed",
    "plan","warranty","case","cover","protector","mount","bracket",
    "hdmi cable","soundbar","grille kit","door panel kit","bezel",
    "frame kit","motorized lift","tv lift","remote only",
]

def _extract_sku(url):
    if not url: return None
    m = re.search(r'skuId=(\d+)|/(\d{7,8})\.p', str(url))
    return (m.group(1) or m.group(2)) if m else None


def make_url(name, sku):
    slug = re.sub(r'[^a-z0-9\s-]', '', name.lower().strip())
    slug = re.sub(r'\s+', '-', slug)
    slug = re.sub(r'-+', '-', slug).strip('-')[:80]
    return f"https://www.bestbuy.com/site/{slug}/{sku}.p?skuId={sku}"

def check_sku(sku):
    """Verify a known SKU directly — fastest and most accurate."""
    url = f'https://api.bestbuy.com/v1/products/{sku}.json'
    params = {'apiKey': BESTBUY_API_KEY,
              'show': 'sku,name,salePrice,regularPrice,onlineAvailability'}
    try:
        r = requests.get(url, params=params, timeout=10)
        if r.status_code == 200:
            d = r.json()
            price = d.get('salePrice') or d.get('regularPrice')
            return price, d.get('onlineAvailability', False)
        return None, False
    except:
        return None, False

def search_bb(term, pmin, pmax, must_contain=None):
    safe = re.sub(r"[^a-zA-Z0-9 .&/-]", " ", term).strip()
    q    = f'(search={requests.utils.quote(safe)}&salePrice>={int(pmin)}&salePrice<={int(pmax)})'
    url  = f'https://api.bestbuy.com/v1/products{q}'
    params = {'apiKey': BESTBUY_API_KEY, 'format': 'json',
              'show': 'sku,name,salePrice,regularPrice,onlineAvailability',
              'pageSize': 8, 'sort': 'bestSellingRank.asc'}
    try:
        r = requests.get(url, params=params, timeout=12)
        if r.status_code != 200: return None
        for p in r.json().get('products', []):
            nl = p.get('name','').lower()
            if any(e in nl for e in EXCLUDE): continue
            if not p.get('onlineAvailability', False): continue
            if must_contain and not any(m.lower() in nl for m in must_contain): continue
            price = p.get('salePrice') or p.get('regularPrice')
            if price: return price
    except:
        pass
    return None

# ── LOAD ENTRIES AND BASELINE PRICES ─────────────────────────────────────────
with open(ENTRIES_FILE, 'rb') as f:
    saved = pickle.load(f)

entries    = saved['entries']
results    = saved['results']
entry_map  = {e['entry_id']: e for e in entries}

# ── CHECK PRICES ──────────────────────────────────────────────────────────────
print(f"\nRunning daily check — {TODAY_LABEL}")
print(f"Entries: {len(entries)} | Manual/skip: {sum(1 for r in results.values() if r['status']=='manual')}")

today_prices = {}  # eid -> current price
unchanged = changed = unavailable = skipped = errors = 0
change_log = []    # for email summary

for e in entries:
    eid      = e['entry_id']
    r        = results.get(eid, {})
    baseline = r.get('bb_price')

    if r.get('status') in ('manual', 'not_found'):
        today_prices[eid] = None
        skipped += 1
        continue

    # Try to get current price — use SKU if we have it
    sku = r.get('sku') or _extract_sku(r.get('bb_url',''))
    if sku:
        current_price, avail = check_sku(sku)
        if not avail and current_price:
            unavailable += 1
    else:
        must = e.get('must_contain', ['tv' if eid.startswith('TV') else 'refrigerator'])
        pmin = e.get('price_low', 0) * 0.85
        pmax = e.get('price_high', 9999) * 1.20
        current_price = search_bb(e.get('search_term',''), pmin, pmax, must)

    today_prices[eid] = current_price
    time.sleep(0.35)

    if current_price is None:
        errors += 1
        continue

    if baseline:
        diff     = current_price - baseline
        diff_pct = (diff / baseline) * 100
        if abs(diff_pct) >= 1.0:
            changed += 1
            direction = "📈 UP" if diff > 0 else "📉 DOWN"
            change_log.append({
                'eid':      eid,
                'name':     r.get('bb_name','')[:55],
                'baseline': baseline,
                'current':  current_price,
                'diff':     diff,
                'pct':      diff_pct,
                'dir':      direction,
            })
            print(f"  {direction} {eid} ${baseline:.2f}→${current_price:.2f} ({diff_pct:+.1f}%)")
        else:
            unchanged += 1
    else:
        unchanged += 1

# ── UPDATE EXCEL ──────────────────────────────────────────────────────────────
print(f"\nUpdating {EXCEL_FILE}...")
wb = load_workbook(EXCEL_FILE)

thin = Side(style='thin', color='D1D5DB')
T    = Border(left=thin, right=thin, top=thin, bottom=thin)
TEAL_L  = "E0FAF5"; LGRAY = "F3F4F6"; DGRAY = "374151"
GREEN_BG= "DCFCE7"; GREEN_FG="15803D"
RED_BG  = "FEE2E2"; RED_FG ="DC2626"
AMBER_L = "FFFBEB"; AMBER  ="F59E0B"
NAVY    = "0D1B2A"; WHITE  ="FFFFFF"


for ws in wb.worksheets:
    if ws.title.startswith('📊'): continue  # skip summary tab

    # Find the header row (row 5) and last used price column
    header_row = 5
    last_col   = ws.max_column

    # Add new date column header
    new_col = last_col + 1
    ws.column_dimensions[ws.cell(row=1, column=new_col).column_letter].width = 12
    hdr = ws.cell(row=header_row, column=new_col, value=TODAY_LABEL)
    hdr.font      = Font(color=WHITE, bold=True, size=9, name='Arial')
    hdr.fill      = PatternFill('solid', start_color="15803D")
    hdr.alignment = Alignment(horizontal='center', vertical='center')
    hdr.border    = T

    # Determine which category this sheet is
    is_tv = '📺' in ws.title

    # Map entry IDs to row numbers by scanning column B (entry #) and M (BB name)
    # Find data rows — look for rows where col N (Entered Price) has a value
    for data_row in range(6, ws.max_row + 1):
        entered_cell = ws.cell(row=data_row, column=14)  # col N = Entered Price
        bb_name_cell = ws.cell(row=data_row, column=13)  # col M = BB Name

        if not entered_cell.value: continue
        if not isinstance(entered_cell.value, (int, float)): continue

        # Match this row to an entry by finding the entry whose BB price matches
        baseline_price = float(entered_cell.value)
        matched_eid    = None

        for eid, r in results.items():
            if r.get('bb_price') and abs(r['bb_price'] - baseline_price) < 0.02:
                cat_match = (eid.startswith('TV') == is_tv)
                if cat_match:
                    matched_eid = eid
                    break

        if not matched_eid: continue

        current = today_prices.get(matched_eid)
        baseline = baseline_price

        if current is None:
            cell = ws.cell(row=data_row, column=new_col, value='—')
            cell.font      = Font(color=DGRAY, size=9, name='Arial')
            cell.fill      = PatternFill('solid', start_color=LGRAY)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = T
        else:
            diff_pct = ((current - baseline) / baseline * 100) if baseline else 0
            if abs(diff_pct) < 1:
                bg, fg = TEAL_L, "00897B"
            elif current > baseline:
                bg, fg = AMBER_L, AMBER
            else:
                bg, fg = GREEN_BG, GREEN_FG

            cell = ws.cell(row=data_row, column=new_col, value=current)
            cell.font         = Font(color=fg, bold=(abs(diff_pct)>=1), size=10, name='Arial')
            cell.fill         = PatternFill('solid', start_color=bg)
            cell.number_format = '$#,##0.00'
            cell.alignment    = Alignment(horizontal='center', vertical='center')
            cell.border       = T

wb.save(EXCEL_FILE)
print(f"Excel updated — new column added: {TODAY_LABEL}")

# ── SEND EMAIL ────────────────────────────────────────────────────────────────
total_checked = unchanged + changed + unavailable + errors
subject = f"📊 Price Book Update — {TODAY_LABEL} | {changed} changes found"

# Build HTML email
rows_html = ""
if change_log:
    for c in sorted(change_log, key=lambda x: abs(x['pct']), reverse=True):
        color   = "#DC2626" if c['diff'] > 0 else "#15803D"
        bg      = "#FFF7ED" if c['diff'] > 0 else "#DCFCE7"
        sku  = results.get(c['eid'],{}).get('sku','—')
        burl = results.get(c['eid'],{}).get('bb_url','')
        sku_link = f'<a href="{burl}" style="color:#2563EB;font-size:10px">{sku}</a>' if burl else sku
        rows_html += f"""
        <tr style="background:{bg}">
          <td style="padding:6px 10px;font-size:12px;color:#374151">{c['eid']}</td>
          <td style="padding:6px 10px;font-size:12px;color:#374151">{c['name']}</td>
          <td style="padding:6px 10px;font-size:12px;text-align:center;color:#6B7280">{sku_link}</td>
          <td style="padding:6px 10px;font-size:12px;text-align:center">${c['baseline']:.2f}</td>
          <td style="padding:6px 10px;font-size:12px;text-align:center;font-weight:bold;color:{color}">${c['current']:.2f}</td>
          <td style="padding:6px 10px;font-size:12px;text-align:center;font-weight:bold;color:{color}">{c['dir']} {c['pct']:+.1f}%</td>
        </tr>"""
else:
    rows_html = '<tr><td colspan="5" style="padding:12px;text-align:center;color:#6B7280">No price changes detected today</td></tr>'

html = f"""
<div style="font-family:Arial,sans-serif;max-width:700px;margin:0 auto">
  <div style="background:#0D1B2A;padding:20px;border-radius:8px 8px 0 0">
    <h2 style="color:white;margin:0">📊 Daily Price Book Update</h2>
    <p style="color:#9CA3AF;margin:4px 0 0">{date.today().strftime('%B %d, %Y')}</p>
  </div>
  <div style="background:#F9FAFB;padding:20px;border:1px solid #E5E7EB">
    <div style="display:flex;gap:16px;margin-bottom:20px">
      <div style="flex:1;background:white;border:1px solid #E5E7EB;border-radius:6px;padding:12px;text-align:center">
        <div style="font-size:24px;font-weight:bold;color:#0D1B2A">{total_checked}</div>
        <div style="font-size:11px;color:#6B7280">Checked</div>
      </div>
      <div style="flex:1;background:#DCFCE7;border:1px solid #BBF7D0;border-radius:6px;padding:12px;text-align:center">
        <div style="font-size:24px;font-weight:bold;color:#15803D">{unchanged}</div>
        <div style="font-size:11px;color:#15803D">Unchanged</div>
      </div>
      <div style="flex:1;background:#{'FFF7ED' if changed else 'F9FAFB'};border:1px solid #{'FED7AA' if changed else '#E5E7EB'};border-radius:6px;padding:12px;text-align:center">
        <div style="font-size:24px;font-weight:bold;color:#{'EA580C' if changed else '#6B7280'}">{changed}</div>
        <div style="font-size:11px;color:#{'EA580C' if changed else '#6B7280'}">Changed</div>
      </div>
      <div style="flex:1;background:#F9FAFB;border:1px solid #E5E7EB;border-radius:6px;padding:12px;text-align:center">
        <div style="font-size:24px;font-weight:bold;color:#6B7280">{skipped}</div>
        <div style="font-size:11px;color:#6B7280">Manual/Skip</div>
      </div>
    </div>
    <table style="width:100%;border-collapse:collapse;background:white;border:1px solid #E5E7EB;border-radius:6px;overflow:hidden">
      <thead>
        <tr style="background:#1F2937">
          <th style="padding:8px 10px;text-align:left;color:white;font-size:11px">ID</th>
          <th style="padding:8px 10px;text-align:left;color:white;font-size:11px">Item</th>
          <th style="padding:8px 10px;text-align:center;color:white;font-size:11px">SKU</th>
          <th style="padding:8px 10px;text-align:center;color:white;font-size:11px">Baseline</th>
          <th style="padding:8px 10px;text-align:center;color:white;font-size:11px">Today</th>
          <th style="padding:8px 10px;text-align:center;color:white;font-size:11px">Change</th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>
    <p style="font-size:11px;color:#9CA3AF;margin-top:16px">
      PriceBook_Combined_Final.xlsx updated — column {TODAY_LABEL} added. &nbsp;
      <a href="https://github.com/joeybukowski3/PRICEBOOK/raw/main/PriceBook_Combined_Final.xlsx" 
         style="color:#2563EB;font-weight:bold">⬇ Download Latest Excel</a>
      &nbsp;|&nbsp;
      <a href="https://github.com/joeybukowski3/PRICEBOOK" style="color:#2563EB">View Repo</a>
    </p>
  </div>
</div>
"""

payload = {
    "from":    "PriceBook <onboarding@resend.dev>",
    "to":      [NOTIFY_EMAIL],
    "subject": subject,
    "html":    html,
}
resp = requests.post(
    "https://api.resend.com/emails",
    headers={"Authorization": f"Bearer {RESEND_API_KEY}",
             "Content-Type": "application/json"},
    data=json.dumps(payload),
    timeout=15,
)
if resp.status_code in (200, 201):
    print(f"Email sent to {NOTIFY_EMAIL}")
else:
    print(f"Email failed: {resp.status_code} {resp.text}")

print(f"\nDone — {unchanged} unchanged | {changed} changed | {skipped} skipped")


