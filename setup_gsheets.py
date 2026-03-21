"""
One-time setup: migrates GITHUBOPENCLAWPRICECHECK.xlsx into Google Sheets.
Run once from your PRICEBOOK folder before switching to openclaw_gsheets.py.
"""
import os, json, time
import openpyxl
import gspread
from google.oauth2.service_account import Credentials

SHEET_ID = '1-i5CuSAXKFrvULPn_9bI5g5yCI3djy_AwC3c0apgkLY'
SCOPES   = ['https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive']

def get_client():
    with open('gsheet_credentials.json') as f:
        info = json.load(f)
    creds = Credentials.from_service_account_info(info, scopes=SCOPES)
    return gspread.authorize(creds)

print("Connecting to Google Sheets...")
gc          = get_client()
spreadsheet = gc.open_by_key(SHEET_ID)
print(f"Connected: {spreadsheet.title}")

# Read Excel
print("Reading GITHUBOPENCLAWPRICECHECK.xlsx...")
wb = openpyxl.load_workbook('GITHUBOPENCLAWPRICECHECK.xlsx')
ws = wb.active

# Find last real row
last_row = 1
for row in range(ws.max_row, 1, -1):
    if ws.cell(row=row, column=3).value:
        last_row = row
        break

# Read all data
all_data = []
for row in range(1, last_row + 1):
    row_data = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=row, column=col).value
        if hasattr(val, 'strftime'):
            val = f"{val.month}/{val.day}"
        elif val is None:
            val = ''
        else:
            val = str(val) if not isinstance(val, (int, float)) else val
        row_data.append(val)
    all_data.append(row_data)

print(f"Read {len(all_data)} rows, {len(all_data[0]) if all_data else 0} columns")

# Get or create OPENCLAW tab
try:
    ws_oc = spreadsheet.worksheet("OPENCLAW")
    ws_oc.clear()
    print("Cleared existing OPENCLAW tab")
except gspread.exceptions.WorksheetNotFound:
    ws_oc = spreadsheet.add_worksheet(title="OPENCLAW", rows=200, cols=50)
    print("Created OPENCLAW tab")

# Write data in chunks
print("Writing data to Google Sheets...")
chunk_size = 50
for i in range(0, len(all_data), chunk_size):
    chunk = all_data[i:i+chunk_size]
    start_row = i + 1
    end_row   = start_row + len(chunk) - 1
    cols      = len(chunk[0]) if chunk else 1

    def col_letter(n):
        result = ""
        while n:
            n, r = divmod(n-1, 26)
            result = chr(65+r) + result
        return result

    range_name = f"A{start_row}:{col_letter(cols)}{end_row}"
    ws_oc.update(range_name, chunk, value_input_option='RAW')
    time.sleep(0.5)
    print(f"  Wrote rows {start_row}–{end_row}")

print(f"\nSetup complete!")
print(f"Open your sheet: https://docs.google.com/spreadsheets/d/{SHEET_ID}")
print("Now you can run openclaw_gsheets.py for daily checks.")
